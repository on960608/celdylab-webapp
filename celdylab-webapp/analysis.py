"""
시딩 인사이트 / 공구 성과 / 리스트업 판정에서 공통으로 쓰는 계산식.
기존 오프라인 트래커(seeding-gongu-tracker.html)의 자바스크립트 공식을 그대로 옮겼어요.
"""
from datetime import date, datetime

BRANDS = ["코드니처", "빠이러스", "라이프스타일마트"]
TREND_PLATFORMS = ["82market", "지금하는공구", "공구모아", "공구팡팡", "맘캘린더"]
TREND_CATEGORIES = ["리빙", "여행", "홈인테리어", "패션잡화", "주방용품", "생활용품", "기타"]

# 인플루언서 공동구매(공구)가 활발히 진행되는 모니터링 대상 플랫폼 — 5개 모두 자동 수집돼요
# (trend_scraper.py 참고). 원래는 캘린·위시버니·인공도 모니터링했지만 화면 구조상 자동
# 수집이 불가능해서(자바스크립트로만 목록을 불러오거나 비공개 API 방식) 제외했고, 그 세
# 자리는 SSR로 확인 가능한 공구팡팡(09pangpang.com)·맘캘린더(momcalendar.com) 두 곳으로
# 대체했어요 — 나머지 한 자리를 대신할 만한 사이트는 끝내 찾지 못해서(대부분 비공개 API
# 방식이거나 이 장르가 아닌 B2B 벤더/마케팅 플랫폼) 지금은 5개로 운영해요.
# 공구모아는 원래 gonggumoa.com이었지만 그 사이트는 자바스크립트로만 목록을 불러와
# 자동화가 불가능해서, 같은 이름을 쓰는 09more.com(자동 수집 가능)으로 대체했어요.
TREND_PLATFORM_LINKS = [
    {"name": "82market", "url": "https://www.82market.com/", "desc": "인플루언서 공구 마켓", "auto": True},
    {"name": "지금하는공구", "url": "https://www.09now.com/", "desc": "인스타 공구 검색엔진", "auto": True},
    {"name": "공구모아", "url": "https://www.09more.com/", "desc": "SNS 공동구매 정보 통합 모음", "auto": True},
    {"name": "공구팡팡", "url": "https://09pangpang.com/", "desc": "인스타 최신 공구 모음", "auto": True},
    {"name": "맘캘린더", "url": "https://momcalendar.com/", "desc": "인스타 공구 일정 캘린더", "auto": True},
]


# ---------------------------------------------------------------------------
# 시딩 인사이트
# ---------------------------------------------------------------------------

def insight_metrics(r):
    """r: dict-like (views, likes, comments, saves, shares, followers)"""
    views = r["views"] or 0
    followers = r["followers"] or 0
    engagement = ((r["likes"] or 0) + (r["comments"] or 0) + (r["saves"] or 0) + (r["shares"] or 0)) / views * 100 if views else 0
    save_rate = (r["saves"] or 0) / views * 100 if views else 0
    reach_rate = views / followers * 100 if followers else 0
    return {"engagement": engagement, "save_rate": save_rate, "reach_rate": reach_rate}


def pct(n):
    return f"{(n or 0):.1f}%"


def won(n):
    return f"{round(n or 0):,}원"


# ---------------------------------------------------------------------------
# 협찬 인원 리스트업 — STEP 02 · 5가지 탈락 기준 + 반응점수(1,000점)
# ---------------------------------------------------------------------------

def listup_score(c):
    avg_views = ((c["views1"] or 0) + (c["views2"] or 0) + (c["views3"] or 0)) / 3
    return round(avg_views / 10 + (c["comments"] or 0) * 2 + (c["shares"] or 0) * 4)


def days_since(date_str):
    if not date_str:
        return None
    try:
        d = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    return (date.today() - d).days


def listup_exclusion_reasons(c):
    reasons = []
    avg_views = ((c["views1"] or 0) + (c["views2"] or 0) + (c["views3"] or 0)) / 3
    if 0 < avg_views < 10000:
        reasons.append(f"① 최근 릴스 조회수 낮음(평균 {round(avg_views):,})")
    if not c["has_real_comments"]:
        reasons.append("② 실제 관심 댓글 없음")
    likes = c["likes"] or 0
    if likes and ((c["comments"] or 0) + (c["shares"] or 0)) / likes * 100 < 1:
        reasons.append("③ 좋아요 대비 반응 낮음")
    if c["sponsored_low"]:
        reasons.append("④ 협찬 콘텐츠 반응 낮음")
    d = days_since(c["last_upload"])
    if d is not None and d >= 14:
        reasons.append(f"⑤ 최근 활동 없음({d}일 경과)")
    return reasons


def listup_verdict(c):
    score = listup_score(c)
    reasons = listup_exclusion_reasons(c)
    if score < 1000:
        reasons = [f"반응점수 미달({score:,}점)"] + reasons
    if not reasons:
        return {"pass": True, "exception": False, "reasons": [], "score": score}
    if c["reason"]:
        return {"pass": True, "exception": True, "reasons": reasons, "score": score}
    return {"pass": False, "exception": False, "reasons": reasons, "score": score}


# ---------------------------------------------------------------------------
# 공구 성과 분석
# ---------------------------------------------------------------------------

def gongu_net_sold(r):
    return max(0, (r["sold_qty"] or 0) - (r["return_qty"] or 0))


def gongu_return_pct(r):
    return (r["return_qty"] or 0) / r["sold_qty"] * 100 if r["sold_qty"] else 0


def gongu_per1k(r):
    return (r["revenue"] or 0) / r["followers"] * 1000 if r["followers"] else 0


def gongu_tier(followers):
    n = followers or 0
    if n < 10000:
        return "1만 미만"
    if n < 30000:
        return "1만~3만"
    if n < 50000:
        return "3만~5만"
    if n < 100000:
        return "5만~10만"
    if n < 300000:
        return "10만~30만"
    return "30만 이상"


TIER_ORDER = ["1만 미만", "1만~3만", "3만~5만", "5만~10만", "10만~30만", "30만 이상"]


# ---------------------------------------------------------------------------
# 댓글 이벤트 추첨
# ---------------------------------------------------------------------------
import random
import re


def extract_ig_shortcode(url):
    """인스타그램 게시물/릴스 URL에서 짧은 코드를 추출해요. 못 찾으면 None."""
    m = re.search(r"instagram\.com/(?:p|reel|reels|tv)/([A-Za-z0-9_-]+)", url or "")
    return m.group(1) if m else None


def parse_comments_text(raw_text):
    """
    한 줄에 '아이디: 댓글내용' 형식의 텍스트를 댓글 목록으로 변환해요.
    콜론이 없으면 그 줄 전체를 아이디로, 댓글내용은 빈 문자열로 처리해요.
    """
    comments = []
    for line in (raw_text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if ":" in line:
            username, text = line.split(":", 1)
        elif "：" in line:  # 전각 콜론도 지원
            username, text = line.split("：", 1)
        else:
            username, text = line, ""
        username = username.strip().lstrip("@")
        text = text.strip()
        if username:
            comments.append({"username": username, "text": text})
    return comments


_USERNAME_HEADER_KEYS = ["username", "아이디", "계정", "인스타그램아이디", "인스타아이디", "id", "instagram"]
_COMMENT_HEADER_KEYS = ["comment", "댓글", "댓글내용", "text", "content", "내용"]


def _norm_header(h):
    return str(h or "").strip().replace(" ", "").lower()


def _find_col(headers, keys):
    for i, h in enumerate(headers):
        nh = _norm_header(h)
        if not nh:
            continue
        for k in keys:
            nk = _norm_header(k)
            if nh == nk or nk in nh or nh in nk:
                return i
    return None


def parse_comments_spreadsheet(file_storage):
    """
    업로드된 .csv 또는 .xlsx 파일에서 '아이디'/'댓글' 열을 찾아 댓글 목록으로 변환해요.
    (username, comment) 두 열 이름은 유사한 표현(아이디/계정/username, 댓글/comment/내용 등)을 자동으로 인식해요.
    반환: (comments, error)
    """
    filename = (file_storage.filename or "").lower()
    rows = []
    try:
        if filename.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
        else:
            import csv
            import io
            raw = file_storage.read()
            for enc in ("utf-8-sig", "cp949", "utf-8"):
                try:
                    text = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return None, "파일 인코딩을 인식하지 못했어요. UTF-8 또는 CSV(쉼표 구분) 형식으로 저장해 주세요."
            reader = csv.reader(io.StringIO(text))
            rows = [row for row in reader]
    except Exception as e:
        return None, f"파일을 읽는 중 오류가 발생했어요: {e}"

    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return None, "빈 파일이에요."

    headers = [str(c) if c is not None else "" for c in rows[0]]
    uname_idx = _find_col(headers, _USERNAME_HEADER_KEYS)
    comment_idx = _find_col(headers, _COMMENT_HEADER_KEYS)

    if uname_idx is None:
        # 헤더를 못 찾으면 1열=아이디, 2열=댓글로 간주(헤더 없이 바로 데이터가 시작하는 경우)
        uname_idx, comment_idx = 0, (1 if len(headers) > 1 else None)
        data_rows = rows
    else:
        data_rows = rows[1:]

    comments = []
    for r in data_rows:
        if uname_idx >= len(r):
            continue
        username = str(r[uname_idx] or "").strip().lstrip("@")
        if not username:
            continue
        text = str(r[comment_idx]).strip() if (comment_idx is not None and comment_idx < len(r) and r[comment_idx] is not None) else ""
        comments.append({"username": username, "text": text})

    if not comments:
        return None, "파일에서 아이디 열을 찾지 못했어요. 열 이름을 '아이디'/'username'과 '댓글'/'comment'로 맞춰 주세요."
    return comments, None


def draw_giveaway_winners(comments, event_type, keyword, winner_count, excluded_raw):
    """
    반환: (winners, stats)
    winners: [{rank, username, comment, keyword_matched}]
    stats: {total_comments, matched_accounts, final_pool_count, shortage: bool}
    """
    excluded = {
        e.strip().lower().lstrip("@")
        for e in re.split(r"[,\n]+", excluded_raw or "")
        if e.strip()
    }

    total_comments = len(comments)

    # 계정별로 묶기 (동일 계정 여러 댓글 -> 1명)
    by_user = {}
    for c in comments:
        key = c["username"].lower()
        by_user.setdefault(key, {"username": c["username"], "texts": []})
        by_user[key]["texts"].append(c["text"])

    # 1) 조건(키워드) 충족 계정 추리기 — 제외 계정 필터 전
    matched = []
    for key, info in by_user.items():
        if event_type == "keyword":
            hit_text = next((t for t in info["texts"] if keyword and keyword in t), None)
            if hit_text is None:
                continue
            matched.append({"key": key, "username": info["username"], "comment": hit_text, "keyword_matched": 1})
        else:
            matched.append({"key": key, "username": info["username"], "comment": info["texts"][0], "keyword_matched": None})

    matched_accounts = len(matched)

    # 2) 제외 계정(본인/브랜드 등) 필터링 -> 최종 추첨 대상
    final_pool = [m for m in matched if m["key"] not in excluded]
    final_pool_count = len(final_pool)

    shortage = final_pool_count < winner_count
    n = min(winner_count, final_pool_count) if winner_count > 0 else 0
    picked = random.sample(final_pool, n) if n > 0 else []
    winners = [
        {"rank": i + 1, "username": w["username"], "comment": w["comment"], "keyword_matched": w["keyword_matched"]}
        for i, w in enumerate(picked)
    ]

    stats = {
        "total_comments": total_comments,
        "matched_accounts": matched_accounts,
        "final_pool_count": final_pool_count,
        "shortage": shortage,
    }
    return winners, stats


def build_winners_excel(event):
    """당첨자 명단(event['winners'])을 엑셀(.xlsx) 파일로 만들어서 메모리 버퍼로 돌려줘요."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "당첨자 명단"

    ws.append(["순번", "Instagram 아이디", "작성 댓글", "필수 단어 포함"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for w in event["winners"]:
        matched = w["keyword_matched"]
        matched_text = "포함" if matched == 1 else ("미포함" if matched == 0 else "-")
        ws.append([w["rank"], w["username"], w["comment_text"] or "", matched_text])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def fetch_ig_comments_via_graph_api(post_url):
    """
    셀디랩이 직접 운영하는 인스타그램 비즈니스 계정에 연동된 게시물의 댓글을 Graph API로 가져와요.
    INSTAGRAM_ACCESS_TOKEN / INSTAGRAM_BUSINESS_ID 환경변수가 없으면 (None, 안내 메시지)를 돌려줘요.
    타사 계정 게시물은 이 방식으로 절대 가져올 수 없어요 (Graph API 자체 제약).
    """
    import os
    import requests

    token = os.environ.get("INSTAGRAM_ACCESS_TOKEN")
    ig_user_id = os.environ.get("INSTAGRAM_BUSINESS_ID")
    if not token or not ig_user_id:
        return None, "INSTAGRAM_ACCESS_TOKEN / INSTAGRAM_BUSINESS_ID 환경변수가 설정되어 있지 않아요. 아래 '댓글 직접 붙여넣기'를 이용해 주세요."

    shortcode = extract_ig_shortcode(post_url)
    if not shortcode:
        return None, "인스타그램 게시물/릴스 링크 형식을 인식하지 못했어요."

    api_version = "v21.0"
    base = f"https://graph.facebook.com/{api_version}"

    # 1) 연동 계정의 미디어 목록에서 permalink가 일치하는 게시물의 media id 찾기
    media_id = None
    url = f"{base}/{ig_user_id}/media"
    params = {"fields": "id,permalink", "access_token": token, "limit": 50}
    for _ in range(20):  # 최대 20페이지(=1000개)까지만 탐색
        try:
            resp = requests.get(url, params=params, timeout=10)
            data = resp.json()
        except Exception as e:
            return None, f"인스타그램 API 요청 중 오류: {e}"
        if "error" in data:
            return None, data["error"].get("message", "Graph API 오류가 발생했어요.")
        for item in data.get("data", []):
            if shortcode in (item.get("permalink") or ""):
                media_id = item["id"]
                break
        if media_id:
            break
        next_url = data.get("paging", {}).get("next")
        if not next_url:
            break
        url, params = next_url, None

    if not media_id:
        return None, "연동된 인스타그램 계정에서 이 링크의 게시물을 찾지 못했어요 (다른 계정의 게시물이거나, 최근 미디어 목록 범위 밖일 수 있어요)."

    # 2) 해당 게시물의 댓글 전체(페이지네이션) 가져오기
    comments = []
    url = f"{base}/{media_id}/comments"
    params = {"fields": "username,text,timestamp", "access_token": token, "limit": 100}
    for _ in range(50):  # 최대 5,000개 댓글까지
        try:
            resp = requests.get(url, params=params, timeout=10)
            data = resp.json()
        except Exception as e:
            return None, f"댓글 조회 중 오류: {e}"
        if "error" in data:
            return None, data["error"].get("message", "Graph API 오류가 발생했어요.")
        for c in data.get("data", []):
            comments.append({"username": c.get("username", ""), "text": c.get("text", "")})
        next_url = data.get("paging", {}).get("next")
        if not next_url:
            break
        url, params = next_url, None

    return comments, None


# ---------------------------------------------------------------------------
# 인스타그램 이벤트 댓글 자동 엑셀 추출
# ---------------------------------------------------------------------------
#
# 인스타그램은 로그인한 사용자의 브라우저에만 댓글 전체를 내려주고(비공개 내부 API +
# 로그인 세션 쿠키 필요), 외부 웹사이트가 게시물 링크만으로 그 내부 API를 서버 간
# 요청으로 직접 호출하는 건 인스타그램의 인증·CORS·봇 차단 정책상 막혀 있어요. 그래서
# "게시물 링크만 입력하면 서버가 알아서 끝까지 다 가져오는" 자동화는 기술적으로
# 불가능해요 — 이 사실을 감춘 채 일부만 가져와지는 걸 마치 전체가 된 것처럼 보여주지
# 않기 위해, 대신 사용자가 이미 성공적으로 써봤던 방법을 그대로 자동화했어요: 브라우저
# 개발자도구 Network 탭에서 댓글을 끝까지 스크롤해 전부 불러온 뒤 "Save all as HAR"로
# 저장한 .har 파일을 올리면, 그 안에 담긴 모든 xdt_api_v1_media_media_id_comments_connection
# (+ 구버전 GraphQL의 edge_media_to_parent_comment 등) 페이지네이션 응답을 자동으로 찾아
# 모으고 댓글 ID 기준으로 중복 제거해서 엑셀로 만들어줘요. 사용자가 직접 해야 하는 유일한
# 수작업은 "HAR 파일 저장"뿐이고, 그 안의 JSON을 일일이 열어보고 파싱하던 수작업은 이제
# 이 서버가 대신해요.


def load_har_entries(raw_bytes):
    """
    업로드된 .har 파일 bytes를 파싱해서 네트워크 요청 entries 목록을 돌려줘요.
    반환: (entries, error)
    """
    import json

    try:
        har = json.loads(raw_bytes)
    except Exception:
        return None, "HAR 파일 형식을 인식하지 못했어요. 개발자도구 Network 탭에서 저장한 .har 파일이 맞는지 확인해 주세요."

    entries = (((har or {}).get("log") or {}).get("entries")) or []
    if not entries:
        return None, "HAR 파일 안에 네트워크 요청 기록이 없어요."
    return entries, None


def _looks_like_comment_node(d):
    """딕셔너리 하나가 '댓글 하나'처럼 생겼는지 판별해요 (text + 작성자 username + id 조합)."""
    if not isinstance(d, dict):
        return False
    if not isinstance(d.get("text"), str):
        return False
    if not any(k in d for k in ("pk", "id", "comment_id", "cid")):
        return False
    for uk in ("user", "owner"):
        u = d.get(uk)
        if isinstance(u, dict) and isinstance(u.get("username"), str) and u.get("username"):
            return True
    return False


def _extract_comment_fields(d):
    comment_id = str(d.get("pk") or d.get("id") or d.get("comment_id") or d.get("cid") or "").strip()
    username = ""
    for uk in ("user", "owner"):
        u = d.get(uk)
        if isinstance(u, dict) and u.get("username"):
            username = u.get("username")
            break
    created_at = d.get("created_at") or d.get("created_at_utc") or d.get("created_time")
    like_count = d.get("comment_like_count")
    if like_count is None:
        like_count = d.get("like_count")
    if like_count is None:
        elb = d.get("edge_liked_by")
        if isinstance(elb, dict):
            like_count = elb.get("count")
    if not isinstance(like_count, int):
        like_count = None
    return {
        "comment_id": comment_id,
        "username": username,
        "text": d.get("text") or "",
        "created_at": created_at,
        "like_count": like_count,
    }


def _walk_for_comments(obj, found, counter, in_comment_ctx=False):
    """
    JSON 트리를 재귀적으로 훑으면서, 키 이름에 'comment'가 들어있는 구간(예:
    xdt_api_v1_media_media_id_comments_connection, edge_media_to_parent_comment,
    edge_threaded_comments 등) 안에서만 댓글 노드를 인식해요 — 캡션처럼 생김새가
    비슷한 다른 데이터를 댓글로 잘못 인식하지 않기 위한 안전장치예요.
    found: {comment_id: comment_dict} — 댓글 ID 기준으로 자동 중복 제거됨.
    counter[0]: 중복 포함, 발견한 댓글 노드 총 개수(나중에 "중복 제거 개수" 계산용).
    """
    if isinstance(obj, dict):
        if in_comment_ctx and _looks_like_comment_node(obj):
            f = _extract_comment_fields(obj)
            if f["comment_id"]:
                counter[0] += 1
                found[f["comment_id"]] = f
        for k, v in obj.items():
            child_ctx = in_comment_ctx or ("comment" in str(k).lower())
            _walk_for_comments(v, found, counter, child_ctx)
    elif isinstance(obj, list):
        for v in obj:
            _walk_for_comments(v, found, counter, in_comment_ctx)


def extract_comments_from_har_entry(entry, found, counter):
    """
    HAR entry(네트워크 응답) 하나에서 댓글 데이터를 찾아 found에 누적해요.
    이 응답에서 새로 찾은 댓글이 하나라도 있으면 True를 돌려줘요.
    """
    import base64
    import json

    try:
        content = ((entry.get("response") or {}).get("content") or {})
        text = content.get("text")
        if not text:
            return False
        if content.get("encoding") == "base64":
            try:
                text = base64.b64decode(text).decode("utf-8", errors="ignore")
            except Exception:
                return False
        text = text.strip()
        if not text or text[0] not in "{[":
            return False
        data = json.loads(text)
    except Exception:
        return False

    before = len(found)
    _walk_for_comments(data, found, counter, False)
    return len(found) > before


def build_ig_comments_excel(comments):
    """
    댓글 목록을 엑셀(.xlsx)로 만들어서 메모리 버퍼로 돌려줘요.
    열: 번호 / Instagram ID / 댓글 내용 / 댓글 ID / 작성 시각 / 좋아요 수.
    1행은 고정(freeze) + 자동 필터가 걸려 있고, 줄바꿈·이모지·한글도 그대로 저장돼요.
    """
    import io
    from datetime import datetime, timezone, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    def fmt_ts(value):
        if value in (None, ""):
            return ""
        try:
            ts = float(value)
        except (TypeError, ValueError):
            return str(value)
        try:
            dt = datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(timezone(timedelta(hours=9)))
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)

    wb = Workbook()
    ws = wb.active
    ws.title = "댓글"

    headers = ["번호", "Instagram ID", "댓글 내용", "댓글 ID", "작성 시각", "좋아요 수"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"

    for i, c in enumerate(comments, start=1):
        ws.append([
            i,
            c.get("username") or "",
            c.get("text") or "",
            c.get("comment_id") or "",
            fmt_ts(c.get("created_at")),
            c.get("like_count") if c.get("like_count") is not None else "",
        ])

    for idx, width in enumerate([8, 22, 60, 20, 20, 12], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
